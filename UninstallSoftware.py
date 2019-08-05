from PyQt5.QtWidgets import *
from PyQt5.QtCore import  QCoreApplication, Qt, pyqtSlot,QTranslator
from PyQt5.QtGui import QPixmap,QIcon,QFont
from PyQt5 import QtCore
import win32ui
import win32gui
import win32con
import win32api
import win32process
import psutil
import re
from PIL import Image
from openpyxl.styles import Font, Alignment
import socket,openpyxl,sys,os
from SoftwareTool import Ui_MainWindow
from winregeditor import winregeditor


class ShowWindow(Ui_MainWindow, winregeditor,QMainWindow):
    def __init__(self, parent=None):
        super(ShowWindow, self).__init__(parent)
        self.setupUi(self)
        self.getwinreg()
        self.initUI()
        self.init_software()
 

    def initUI(self):
        self.actionsave.triggered.connect(self.tosaveFile)
        self.actionexit.triggered.connect(QCoreApplication.instance().quit)
        #self.actionexit.triggered.connect(self.close)
        self.lineEdit.returnPressed.connect(self.searchSofeware)
        self.actionrefresh.triggered.connect(self.refresh)
        self.actionuninstall.triggered.connect(self.uninstall)
        self.actionwinregLocation.triggered.connect(self.winregLocation)
        self.actionfolderLocation.triggered.connect(self.folderLocation)
        self.action_english.triggered.connect(self._trigger_english)
        self.action_chinese.triggered.connect(self._trigger_chinese)

        #翻译家
        self.trans = QTranslator()
       

    def init_software(self):
        winrege= winregeditor()
        self.numreg=winrege.getwinreg()
        self.init_tableWidget()

    def init_tableWidget(self):
        self.tableWidget.setRowCount(len(self.numreg[1]))                            
        self.tableWidget.setColumnCount(3)

        #QTableWidget背景色交替
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.setStyleSheet("border 0px; color: #6b6d7b; alternate-background-color: red; background: white;")

        self.label.setText('总共'+str(self.tableWidget.rowCount())+'个软件')
        self.label.setFont(QFont("Roman times",10,QFont.Bold))
       

        #print(self.tableWidget.rowCount())                          
        #print(self.tableWidget.columnCount())

        #设置列宽度，行宽度
        #self.tableWidget.setColumnWidth(0, 1000)                      
        #self.tableWidget.setRowHeight(0, 30)
        self.tableWidget.setHorizontalHeaderLabels(['软件名称','大小','待定'])
        
        #向tablewidget添加数据绑定到槽函数
        self.dispalyName_Icon()
        self.software_size()
       

        #设置水平方向表格为自适应的伸缩模式
        #self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        #self.tableWidget.horizontalHeader().setStretchLastSection(True) #最后一行自适应自动拉伸
        self.tableWidget.horizontalHeader().setSectionResizeMode (0, QHeaderView.Stretch) #第一列拉伸自适应

        #表格头的显示与隐藏
        self.tableWidget.verticalHeader().setVisible(False)
        #self.tableWidget.horizontalHeader().setVisible(False)
        
        #去除鼠标点击的选项会出现虚框
        self.tableWidget.setItemDelegate(NoFocusDelegate())

        #表格中不显示分割线
        self.tableWidget.setShowGrid(False)

        # 将表格变为禁止编辑
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)

        #设置表格整行选中
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        #按照软件名称排序
        #Qt.DescendingOrder降序
        #Qt.AscendingOrder升序
        self.tableWidget.sortItems(0,Qt.AscendingOrder)

        #设置表头排序
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.horizontalHeader().sortIndicatorChanged.connect(self.handleSortIndicatorChanged)
        

        
        # 允许右键产生菜单
        self.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        # 将右键菜单绑定到槽函数generateMenu
        self.tableWidget.customContextMenuRequested.connect(self.generateMenu)

    def handleSortIndicatorChanged(self, index, order):
        if index != 0:
            self.tableWidget.horizontalHeader().setSortIndicator(0, self.sortOrder())
        
    _sort_order = QtCore.Qt.AscendingOrder

    def sortOrder(self):
        return self._sort_order

    def sort(self, column, order):
        if column == 0:
            self._sort_order = order
            QtGui.QStandardItemModel.sort(self, column, order)

    def dispalyName_Icon(self):
        i = 0
        for key in self.numreg[1].keys() :
            try:  # ico 来自exe
                ico_x = win32api.GetSystemMetrics(win32con.SM_CXICON)
                ico_y = win32api.GetSystemMetrics(win32con.SM_CYICON)
                large, small = win32gui.ExtractIconEx(self.numreg[1][key]['exe'], 0)
                #exemenu=self.numreg[1][key]['exe']
                useIcon = large[0]
                destroyIcon = small[0]
                win32gui.DestroyIcon(destroyIcon)

                hdc = win32ui.CreateDCFromHandle(win32gui.GetDC(0))
                hbmp = win32ui.CreateBitmap()
                hbmp.CreateCompatibleBitmap(hdc, ico_x, ico_y)
                hdc = hdc.CreateCompatibleDC()

                hdc.SelectObject(hbmp)
                hdc.DrawIcon((0,0), useIcon)
               
                #savePath = "d:/test.bmp"
                #savePath = "d:/"
                #hbmp.SaveBitmapFile(hdc, savePath)

                bmpstr = hbmp.GetBitmapBits(True)
                img = Image.frombuffer(
                    'RGBA',
                    (32,32),
                    bmpstr, 'raw', 'BGRA', 0, 1
                )
                cwd = os.getcwd()
                img.save(cwd + '/icon.png')
                self.pixmap = 'icon.png'
                
            #except Exception as e:  #ico 来自 icon
            except:  
                #ico 来自 icon
                # 判断ico文件是否存在
                if 'icon' in self.numreg[1][key]  and os.path.isfile(self.numreg[1][key]['icon']):   
                    self.pixmap = QPixmap(self.numreg[1][key]['icon'])
                    iconMenu = self.numreg[1][key]['icon']
                    #split = iconMenu.split('\\')
                    #exeMenu ='\\'.join(split[:-1])
                else:  # 不存在ico文件给定默认图标
                    self.pixmap = 'SoftwareTool.png'
                    #exeMenu = ''

            DisplayName=self.numreg[1][key]['DisplayName'].encode('utf-8')
            DisplayName=str(DisplayName,encoding='utf-8')
            newitem = QTableWidgetItem(QIcon(self.pixmap), DisplayName)
            #newitem1 = QTableWidgetItem(filesize)
            #为每个表格内添加数据
            self.tableWidget.setItem(i,0,newitem)
            #self.tableWidget.setItem(i,1,newitem1)
            i+=1

    #tablewidget中右击菜单
    def generateMenu(self, pos):
        # 计算有多少条数据，默认-1,
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()
        # 表格中的有效数据，支持右键弹出菜单
        if row_num < row_num+1:
            menu = self.menu_2
            #self.actionuninstall
            menu.exec_(self.tableWidget.viewport().mapToGlobal(pos))
        

    #软件大小
    def software_size(self):
        i = 0
        for key in self.numreg[1].keys() :
            InstallLocation=self.numreg[1][key]['InstallLocation']
            if InstallLocation == ''and 'exe' in self.numreg[1][key] :
                exemenu = self.numreg[1][key]['exe']
                split = exemenu.split('\\')
                exeMenu ='\\'.join(split[:-1])
                size = self.getdirsize(exeMenu)
                size = self.covertsize(size)
            elif InstallLocation == '':
                size = 0
            else:
                InstallLocation=str(InstallLocation,encoding='utf-8')
                size = self.getdirsize(InstallLocation)
                size = self.covertsize(size)

            newitem1 = QTableWidgetItem(size)
            if type(size)==float:
                newitem1.setData(QtCore.Qt.DisplayRole,size)
            else:
                newitem1.setText(size)
                #为每个表格内添加数据
            self.tableWidget.setItem(i,1,newitem1)
            self.tableWidget.item(i,1).setTextAlignment(Qt.AlignCenter)
            del newitem1
            i+=1

    def getdirsize(self,path):
        size = 0
        if path == '':
            pass
        else:
            if os.path.exists(path):
                for entry in os.scandir(path):
                    if entry.is_file():
                        size += entry.stat().st_size
                    elif entry.is_dir():
                        size += self.getdirsize(entry.path)
                return size
            else:
                pass
        
    def covertsize(self,size):
        if size is not None:
            if size/1024>1:
                if size/(1024*1024*1024)>1:
                    return '%0.1f GB' %float(size/(1024*1024*1024))
                elif size/(1024*1024)>1:
                    return '%0.1f MB' %float(size/(1024*1024))
                else:
                    return '%0.1f KB' %float(size/1024)
            else:
                return '%0.1f KB' %float(size/1024)
        else:
            pass


    def Action(self):
        if self.actionuninstall.isEnabled():
            self.menubar.setEnabled(False)
            self.centralwidget.setEnabled(False)
            

    #软件卸载
    def uninstall(self):
        #self.Action()
        self.tableWidget.setSortingEnabled(False)
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()

        for key in self.numreg[1].keys() :
            UninstallString =''
            DisplayName=self.numreg[1][key]['DisplayName'].encode('utf-8')
            DisplayName=str(DisplayName,encoding='utf-8')
            if DisplayName==self.tableWidget.item(row_num,0).text():
                UninstallString=self.numreg[1][key]['UninstallString']
                UninstallString=str(UninstallString,encoding='utf-8')
                try:
                    win32api.ShellExecute(0, 'open', UninstallString, '', '', 1)
                    '''
                    #handle = win32process.CreateProcess(UninstallString, '', None, None, 0, win32process.CREATE_NO_WINDOW, None, None, win32process.STARTUPINFO()) 
                    p = subprocess.Popen(UninstallString,stdout=subprocess.PIPE)
                    pid = p.pid
                    if pid is None:
                        self.menubar.setEnabled(True)
                        self.centralwidget.setEnabled(True)
                    print(pid)
                    split = UninstallString.split('\\')
                    exe_name ='\\'.join(split[-1:])
                    for p in psutil.process_iter():
                        if p.name()== exe_name:
                            for child in p.children():
                                if child.pid is None:
                                    self.Action()
                                print(child.pid)
                                #os.kill(child.pid,-1)
                            os.kill(p.pid,-1)
                            if p.name is None:
                                self.Action()
                    '''       
                    
                except:
                    os.system(UninstallString)
                
       
        self.refresh()             
               
  

    #软件注册表位置
    def winregLocation(self):
        #判断regedit注册表是否已打开。若打开，则关闭。
        # 获取所有正在运行中的进程 
        list = psutil.pids() 
        # 遍历列表获取进程名'regedit.exe' 
        for i in range(0, len(list)): 
            try: 
                p = psutil.Process(list[i]) 
                if p.cmdline()[0].find("regedit.exe") != -1: 
                    # 杀掉进程regedit
                    p.kill() 
            except: 
                pass 
        #获取tablewidget中行的位置号
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()
        #获取键值
        for key in self.numreg[1].keys() :
            DisplayName=self.numreg[1][key]['DisplayName'].encode('utf-8')
            DisplayName=str(DisplayName,encoding='utf-8')
            if DisplayName==self.tableWidget.item(row_num,0).text():
                KeyPath = self.numreg[1][key]['KeyPath']
                key_pa ="Software\Microsoft\Windows\CurrentVersion\Applets\Regedit"
                key = win32api.RegOpenKey(win32con.HKEY_CURRENT_USER,key_pa,0, win32con.KEY_ALL_ACCESS)
                #给Lastkey的键添加键值
                win32api.RegSetValueEx(key,"LastKey",0,win32con.REG_SZ,KeyPath)
                win32process.CreateProcess('C:\\Windows\\regedit.exe', '', None, None, 0, win32process.CREATE_NO_WINDOW, None, None, win32process.STARTUPINFO()) 
                      
           
    

    #软件安装目录路径
    def folderLocation(self):
        row_num = -1
        for i in self.tableWidget.selectionModel().selection().indexes():
            row_num = i.row()
        #获取键值
        for key in self.numreg[1].keys() :
            InstallLocation=''
            DisplayName=self.numreg[1][key]['DisplayName'].encode('utf-8')
            DisplayName=str(DisplayName,encoding='utf-8')
            if DisplayName==self.tableWidget.item(row_num,0).text():
                InstallLocation=self.numreg[1][key]['InstallLocation']
                if InstallLocation == ''and 'exe' in self.numreg[1][key] :
                    exemenu = self.numreg[1][key]['exe']
                    split = exemenu.split('\\')
                    exeMenu ='\\'.join(split[:-1])
                    win32api.ShellExecute(0, 'open',exeMenu, '', '', 1)
                else:
                    InstallLocation=str(InstallLocation,encoding='utf-8')
                    win32api.ShellExecute(0, 'open',InstallLocation, '', '', 1)
                    #QMessageBox.information(self,'提示','访问的文件夹不存在！')
                return



    def searchSofeware(self):
        search_text = self.lineEdit.text()
        if not search_text:
            QMessageBox.warning(self, "Warning", '请输入需要查询的软件')
        else:
            # 新建tabelwidget模型
            self.m_model = self.tableWidget
            #搜索栏自动补全方法
            m_completer = QCompleter(self.m_model)
            self.lineEdit.setCompleter(m_completer)
            #激活搜索栏
            m_completer.activated[str].connect(self.onUsernameChoosed)

    def onUsernameChoosed(self, name):
        self.lineEdit.setText(name)  # 设置lineEdit的当前显示文字
        # self.lineEdit.text(name)   #读取当前框中所输入的文字

    @pyqtSlot(str)
    def on_lineEdit_textChanged(self, text):
        if text == "":
            self.tableWidget.setSortingEnabled(False)
            self.init_tableWidget()
        else:
            #重新获取数据之前先关闭可排序性，获取到数据之后再开启排序性
            self.tableWidget.setSortingEnabled(False)
            #新建tabelwidget模型
            self.m_model = self.tableWidget
            #删除已有的行
            for k in range(0,self.m_model.rowCount()):
                self.m_model.removeRow(0)

            i=0
            for key in self.numreg[1].keys() :
                try:  # ico 来自exe
                    ico_x = win32api.GetSystemMetrics(win32con.SM_CXICON)
                    ico_y = win32api.GetSystemMetrics(win32con.SM_CYICON)
                    large, small = win32gui.ExtractIconEx(self.numreg[1][key]['exe'], 0)
                    #exeMenu=self.numreg[1][key]['exe']
                    useIcon = large[0]
                    destroyIcon = small[0]
                    win32gui.DestroyIcon(destroyIcon)

                    hdc = win32ui.CreateDCFromHandle(win32gui.GetDC(0))
                    hbmp = win32ui.CreateBitmap()
                    hbmp.CreateCompatibleBitmap(hdc, ico_x, ico_x)
                    hdc = hdc.CreateCompatibleDC()

                    hdc.SelectObject(hbmp)
                    hdc.DrawIcon((0,0), useIcon)
                    #savePath = "d:/test.bmp"
                    #hbmp.SaveBitmapFile(hdc, savePath)

                    bmpstr = hbmp.GetBitmapBits(True)
                    img = Image.frombuffer(
                        'RGBA',
                        (32,32),
                        bmpstr, 'raw', 'BGRA', 0, 1
                    )
                    cwd = os.getcwd()
                    img.save(cwd + '/icon.png')
                    self.pixmap = 'icon.png'
                except:  
                    #ico 来自 icon
                    # 判断ico文件是否存在
                    if 'icon' in self.numreg[1][key]  and os.path.isfile(self.numreg[1][key]['icon']):   
                        self.pixmap = QPixmap(self.numreg[1][key]['icon'])
                        iconMenu = self.numreg[1][key]['icon']
                        #split = iconMenu.split('\\')
                        #exeMenu ='\\'.join(split[:-1])
                    else:  # 不存在ico文件给定默认图标
                        self.pixmap = 'SoftwareTool.png'
                        #exeMenu = ''

                DisplayName=self.numreg[1][key]['DisplayName'].encode('utf-8')
                DisplayName=str(DisplayName,encoding='utf-8')

                InstallLocation=self.numreg[1][key]['InstallLocation']

                if DisplayName.find(text) >= 0 :
                    rowIndex = self.tableWidget.rowCount()
                    self.tableWidget.setRowCount(rowIndex + 1) #总行数增加1

                    self.label.setText('总共'+str(self.tableWidget.rowCount())+'个软件')

                    newitem = QTableWidgetItem(QIcon(self.pixmap), DisplayName)
                    #为每个表格内添加数据
                    self.tableWidget.setItem(i,0,newitem)
                    
                    if InstallLocation == ''and 'exe' in self.numreg[1][key] :
                        exemenu = self.numreg[1][key]['exe']
                        split = exemenu.split('\\')
                        exeMenu ='\\'.join(split[:-1])
                        size = self.getdirsize(exeMenu)
                        size = str(self.covertsize(size))
                    elif InstallLocation == '':
                        size = 0
                    else:
                        InstallLocation=str(InstallLocation,encoding='utf-8')
                        size = self.getdirsize(InstallLocation)
                        size = str(self.covertsize(size))
                    newitem2 = QTableWidgetItem(size)
                    self.tableWidget.setItem(i,1,newitem2)
                    self.tableWidget.item(i,1).setTextAlignment(Qt.AlignCenter)
                    i+=1 
   
                else:
                    pass
            self.tableWidget.sortItems(0,Qt.AscendingOrder)
            self.tableWidget.setSortingEnabled(True)

            

    #刷新操作
    def refresh(self):
        self.tableWidget.setSortingEnabled(False)
        #新建tabelwidget
        self.m_model = self.tableWidget
        #删除已有的行
        for k in range(0,self.m_model.rowCount()):
            self.m_model.removeRow(0)
        winrege= winregeditor()
        winrege.dicList= {}
        self.numreg=winrege.getwinreg()
        self.init_tableWidget()

    #中英文
    def _trigger_english(self):
        #print("[MainWindow] Change to English")
        self.trans.load("en")
        _app = QApplication.instance()  # 获取app实例
        _app.installTranslator(self.trans)
        # 重新翻译界面
        self.retranslateUi(self)
        pass
    
    def _trigger_chinese(self):
        #print("[MainWindow] Change to English")
        self.trans.load("zh_CN")
        _app = QApplication.instance()  # 获取app实例
        _app.installTranslator(self.trans)
        # 重新翻译界面
        self.retranslateUi(self)
        pass

    #导出excel表格       
    def tosaveFile(self):
        self.cwd = os.getcwd()  # 获取当前程序文件位置
        hostname = socket.gethostname()

        # 获取保存文件的路径信息，file_path返回值为str
        file_path, filetype = QFileDialog.getSaveFileName(self,
                                                          "保存为",
                                                          hostname,  # 起始路径
                                                          "Text Files (*.xlsx);;All Files (*)")
        if file_path == "":
            print("\n取消选择")
            return
        else:
            # 将获取到的路径传入toexcel方法
            self.toexcel(file_path)
            QMessageBox.information(self, "导出excel文件", "文件导出成功.")
            print("\n你选择要保存的文件为:")
            print(file_path)
            print("文件筛选器类型: ", filetype)
            return

    # 设置path参数，用于wb.save的文件存储路径
    def toexcel(self, path):
        # 创建excel表单簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Software"
        _cell = ws.cell(row=1, column=1, value='已安装程序列表')
        # Font properties
        _cell.font = Font(name='微软雅黑', size=14, bold=True)  # bold=True加粗
        # 第一行第一列单元格内容水平居中和垂直居中
        _cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设置表单表格宽度
        ws.column_dimensions['A'].width = 90.0
        # 依次写入list数据
        h = winregeditor().getwinreg()
        for l in range(0, len(h[0])):
            ws.cell(row=l + 2, column=1, value=h[0][l])
        wb.save(path)

#表格中不显示分割线
class NoFocusDelegate(QStyledItemDelegate):
    def paint(self, QPainter, QStyleOptionViewItem, QModelIndex):
        if (QStyleOptionViewItem.state & QStyle.State_HasFocus):
            QStyleOptionViewItem.state = QStyleOptionViewItem.state^QStyle.State_HasFocus
        QStyledItemDelegate.paint(self,QPainter, QStyleOptionViewItem, QModelIndex)
